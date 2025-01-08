# coding:utf-8
import os
import cv2
import time
import json
import ddddocr
from natsort import natsorted
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QDateTime
from PyQt6.QtGui import QColor, QPixmap
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidgetItem, QHeaderView, QSpacerItem, QSizePolicy,
    QHBoxLayout, QLabel
)
from qfluentwidgets import (
    TableWidget, PushButton, ProgressRing, InfoBar, InfoBarPosition, TextEdit, MessageBox
)
import numpy as np
import requests
import base64
import urllib.parse  # 导入urllib.parse用于URL编码
from pymediainfo import MediaInfo
from datetime import datetime, timedelta
import pytz
from requests.exceptions import RequestException  # 导入RequestException
import shutil
import traceback
import re


def load_auth_code(log_signal):
    """从config.json文件加载AuthCode"""
    try:
        if not os.path.exists('config.json'):
            log_signal.emit("config.json配置文件获取失败")
            return None

        with open('config.json', 'r', encoding='utf-8') as config_file:
            config_data = json.load(config_file)
            auth_code = config_data.get("Authorization", {}).get("AuthCode", "")
            if not auth_code:
                log_signal.emit("未配置云端OCR授权码")
                return None
            return auth_code
    except (json.JSONDecodeError, KeyError):
        log_signal.emit("未配置云端OCR授权码")
        return None


def get_baidu_credentials(log_signal):
    """从config.json文件加载百度OCR的API_KEY、SECRET_KEY和URL"""
    try:
        if not os.path.exists('config.json'):
            log_signal.emit("config.json配置文件获取失败")
            return None, None, None

        with open('config.json', 'r', encoding='utf-8') as config_file:
            config_data = json.load(config_file)
            api_key = config_data.get("BaiduOCR", {}).get("ApiKey", "")
            secret_key = config_data.get("BaiduOCR", {}).get("SecretKey", "")
            url = config_data.get("BaiduOCR", {}).get("ApiUrl", "")

            if not api_key:
                log_signal.emit("未配置百度OCR的API_KEY")
            if not secret_key:
                log_signal.emit("未配置百度OCR的SECRET_KEY")
            if not url:
                log_signal.emit("未配置百度OCR的ApiUrl")

            if not api_key or not secret_key or not url:
                return None, None, None
            return api_key, secret_key, url
    except (json.JSONDecodeError, KeyError):
        log_signal.emit("读取config.json配置时发生错误")
        return None, None, None


def get_access_token(api_key, secret_key, log_signal):
    """使用API_KEY和SECRET_KEY获取百度OCR的access_token"""
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {"grant_type": "client_credentials", "client_id": api_key, "client_secret": secret_key}
    try:
        response = requests.post(url, params=params)
        response.raise_for_status()  # 如果返回错误状态码，抛出异常
        return str(response.json().get("access_token"))
    except RequestException:
        log_signal.emit("网络连接失败")
        return None


def replace_text_in_result(result, replace_option, replace_text, log_signal=None):
    if replace_option == "替换文本":
        replacements = replace_text.split(",")
        for replacement in replacements:
            parts = replacement.split("=")
            if len(parts) == 2:
                original, new_text = parts
                result = result.replace(original, new_text)
            else:
                if log_signal:
                    log_signal.emit(f"替换文本格式错误: '{replacement}'")
    elif replace_option == "去掉文本":
        remove_texts = replace_text.split(",")
        for text in remove_texts:
            result = result.replace(text, "")
    return result


class ImageProcessingThread(QThread):
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    task_complete_signal = pyqtSignal(str)
    result_signal = pyqtSignal(object)
    image_visualization_signal = pyqtSignal(str, str)

    def __init__(self, task, stop_flag):
        super().__init__()
        self.task = task
        self.stop_flag = stop_flag
        self.ocr_running = False
        self.ocr_model = None

    def run(self):
        try:
            task_name = self.task['task_name']
            imagestorage_dir = os.path.join('Imagestorage', task_name)
            frame_interval = int(self.task['frame_interval'])
            grayscale_option = self.task['grayscale_option']
            image_time_record = self.task.get('image_time_record', "不记录图片时间")  # 获取新的任务参数

            if os.path.exists(imagestorage_dir):
                self._delete_folder(imagestorage_dir)
                self.log_signal.emit(f"已删除 '{task_name}' 旧文件夹")

            os.makedirs(imagestorage_dir)
            self.log_signal.emit(f"开始执行 '{os.path.basename(self.task['file_path'])}' 转换图片")

            file_path = self.task['file_path']
            if not os.path.exists(file_path):
                self.log_signal.emit(f"'{os.path.basename(file_path)}' 不存在，任务停止")
                return

            # 获取定时执行的时间
            schedule_time_str = self.task['schedule_datetime']  # "HH时mm分"

            if schedule_time_str != "无需定时":
                try:
                    schedule_time = datetime.strptime(schedule_time_str, "%H时%M分").time()
                except ValueError:
                    self.log_signal.emit("定时时间格式错误，任务停止")
                    return

                # 获取当前时间
                current_datetime = datetime.now()

                # 构建定时执行的datetime对象
                scheduled_datetime = datetime.combine(current_datetime.date(), schedule_time)
                if scheduled_datetime <= current_datetime:
                    scheduled_datetime += timedelta(days=1)  # 设置为第二天
            else:
                scheduled_datetime = None  # 无需定时

            recognition_image_paths = self.task['recognition_image_path'].split(",")
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                self.log_signal.emit("无法打开视频文件")
                return

            fps = int(cap.get(cv2.CAP_PROP_FPS))
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            duration = total_frames / fps
            total_steps = int(duration // frame_interval) * len(recognition_image_paths) * 2

            step_count = 0

            for img_name in recognition_image_paths:
                if self.stop_flag['stop']:
                    self.log_signal.emit("任务已停止")
                    return
                img_name = img_name.strip()
                folder_name = img_name.split('.')[0]
                folder_path = os.path.join(imagestorage_dir, folder_name)

                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)

                self.log_signal.emit(f"开始截取 '{img_name}' 图片")
                full_path = os.path.join("Identifyimages", task_name, img_name)
                if not os.path.exists(full_path):
                    self.log_signal.emit(f"'{img_name}' 图片不存在，任务停止")
                    return

                json_path = os.path.join("Identifyimages", task_name, "rectangles.json")
                if not os.path.exists(json_path):
                    self.log_signal.emit(f"'{img_name}' 的 rectangles.json 文件不存在，任务停止")
                    return

                with open(json_path, 'r') as json_file:
                    rectangles = json.load(json_file)

                second_count = 0
                while second_count < duration:
                    if self.stop_flag['stop']:
                        self.log_signal.emit("任务已停止")
                        return
                    cap.set(cv2.CAP_PROP_POS_MSEC, second_count * 1000)
                    ret, frame = cap.read()
                    if not ret:
                        break

                    for rect in rectangles:
                        if rect["index"] == recognition_image_paths.index(img_name) + 1:
                            x, y, w, h = rect["left"], rect["top"], rect["width"], rect["height"]
                            crop_image = frame[y:y + h, x:x + w]

                            # 新增处理图片：放大倍数 (scale_factor)、膨胀迭代次数 (dilation_iterations)、膨胀核大小 (kernel_size)
                            scale_factor = 2  # 放大倍数，默认为2倍
                            dilation_iterations = 1  # 膨胀迭代次数，默认为1次
                            kernel_size = (2, 2)  # 膨胀核大小，默认为(2,2)

                            # 放大图片
                            crop_image = cv2.resize(crop_image, None, fx=scale_factor, fy=scale_factor, interpolation=cv2.INTER_LINEAR)

                            # 膨胀图片
                            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, kernel_size)
                            crop_image = cv2.dilate(crop_image, kernel, iterations=dilation_iterations)

                            # 新增锐化操作
                            sharpening_kernel = np.array([[0, -1, 0],
                                                          [-1, 5, -1],
                                                          [0, -1, 0]])
                            crop_image = cv2.filter2D(crop_image, -1, sharpening_kernel)

                            # 新增功能：检查并处理裁剪后的图片
                            crop_image = self.check_and_process_image(crop_image, task_name)

                            if grayscale_option in ["转换灰度图像", "转换灰度图像(自适应二值化)", "转换灰度图像(固定二值化)"]:
                                crop_image = cv2.cvtColor(crop_image, cv2.COLOR_BGR2GRAY)

                            if grayscale_option == "转换灰度图像(自适应二值化)":
                                crop_image = cv2.adaptiveThreshold(
                                    crop_image, 255,
                                    cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                    cv2.THRESH_BINARY,
                                    11, 2
                                )

                            elif grayscale_option == "转换灰度图像(固定二值化)":
                                _, crop_image = cv2.threshold(crop_image, 127, 255, cv2.THRESH_BINARY)

                            frame_name = os.path.join(folder_path, f"{task_name}_{folder_name}_{second_count:02d}.png")
                            cv2.imencode('.png', crop_image)[1].tofile(frame_name)
                            self.log_signal.emit(f"保存图片: {frame_name}")

                            self.image_visualization_signal.emit(task_name, frame_name)

                    second_count += frame_interval
                    step_count += 1
                    self.progress_signal.emit(int((step_count / total_steps) * 100))

                self.log_signal.emit(f"'{img_name}' 截取完毕")

            cap.release()
            self.log_signal.emit("截取图片完毕，正在执行识别初始化...")

            self.ocr_running = True
            self.run_ocr(task_name, recognition_image_paths, total_steps, scheduled_datetime, image_time_record)

            if not self.stop_flag['stop']:
                self.task_complete_signal.emit(task_name)

        except Exception as e:
            self.log_signal.emit(f"处理过程中出现错误: {e}\n{traceback.format_exc()}")

        finally:
            self.release_model()

    def check_and_process_image(self, image, task_name):
        """
        检查裁剪后的图片是否符合要求，不符合则进行处理。
        要求：
        - 文件大小不超过10MB
        - 最短边至少15px
        - 最长边不超过4096px
        """
        # 首先，检查尺寸
        height, width = image.shape[:2]
        min_side = min(height, width)
        max_side = max(height, width)

        # 标记是否需要调整
        needs_resize = False

        # 检查最短边
        if min_side < 15:
            self.log_signal.emit(f"图片 '{task_name}' 最短边 {min_side}px 小于15px，进行调整")
            needs_resize = True

        # 检查最长边
        if max_side > 4096:
            self.log_signal.emit(f"图片 '{task_name}' 最长边 {max_side}px 超过4096px，进行调整")
            needs_resize = True

        if needs_resize:
            # 计算缩放比例
            scale = 1.0
            if max_side > 4096:
                scale = 4096 / max_side
            if min_side * scale < 15:
                scale = 15 / min_side

            new_width = int(width * scale)
            new_height = int(height * scale)

            # 防止尺寸过小或过大
            new_width = max(new_width, 15)
            new_height = max(new_height, 15)
            new_width = min(new_width, 4096)
            new_height = min(new_height, 4096)

            image = cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_AREA)
            self.log_signal.emit(f"图片 '{task_name}' 调整尺寸到 {new_width}x{new_height}px")

        # 检查文件大小
        is_ok = False
        attempt = 0
        max_attempts = 5
        while not is_ok and attempt < max_attempts:
            success, encoded_image = cv2.imencode('.png', image)
            if not success:
                self.log_signal.emit(f"图片 '{task_name}' 编码失败，尝试重新调整")
                # 尝试压缩
                image = self.compress_image(image)
                attempt += 1
                continue

            size_mb = len(encoded_image.tobytes()) / (1024 * 1024)
            if size_mb <= 10:
                is_ok = True
            else:
                self.log_signal.emit(f"图片 '{task_name}' 大小为 {size_mb:.2f}MB，超过10MB，进行压缩")
                image = self.compress_image(image)
                attempt += 1

        if not is_ok:
            self.log_signal.emit(f"图片 '{task_name}' 无法压缩到10MB以下，使用当前大小")

        return image

    def compress_image(self, image):
        """
        压缩图片以减少文件大小。
        这里通过降低分辨率和调整颜色深度来实现压缩。
        """
        # 降低分辨率的一个简单方法是缩小图片
        height, width = image.shape[:2]
        scale = 0.9  # 每次缩小10%
        new_width = int(width * scale)
        new_height = int(height * scale)
        if new_width < 15 or new_height < 15:
            new_width = max(new_width, 15)
            new_height = max(new_height, 15)
        image = cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_AREA)
        self.log_signal.emit(f"图片压缩至 {new_width}x{new_height}px")
        return image

    def run_ocr(self, task_name, recognition_image_paths, total_steps, scheduled_datetime, image_time_record):
        try:
            self.initialize_model()

            model = self.task['model']
            replace_option = self.task['replace_option']
            replace_text = self.task['replace_text']
            imagestorage_dir = os.path.join('Imagestorage', task_name)

            step_count = total_steps // 2
            ocr_results = {}

            # 获取百度或云端OCR的凭证
            if model == "百度OCR(联网)":
                api_key, secret_key, base_url = get_baidu_credentials(self.log_signal)
                if not api_key or not secret_key or not base_url:
                    self.stop_flag['stop'] = True
                    return

                access_token = get_access_token(api_key, secret_key, self.log_signal)
                if not access_token:
                    self.stop_flag['stop'] = True
                    return

                url = base_url + access_token

            elif model == "云端OCR(联网)":
                auth_code = load_auth_code(self.log_signal)
                if not auth_code:
                    self.stop_flag['stop'] = True
                    return
            else:
                auth_code = None  # 本地模型无需授权码

            for img_name in recognition_image_paths:
                if self.stop_flag['stop']:
                    self.log_signal.emit("OCR任务已停止")
                    self.result_signal.emit((task_name, ocr_results))
                    return
                img_name = img_name.strip()
                self.log_signal.emit(f"识别 '{img_name}' 开始")
                full_path = os.path.join("Identifyimages", task_name, img_name)
                if not os.path.exists(full_path):
                    self.log_signal.emit(f"'{img_name}' 图片不存在，任务停止")
                    self.result_signal.emit((task_name, ocr_results))
                    return

                folder_name = img_name.split('.')[0]
                folder_path = os.path.join(imagestorage_dir, folder_name)
                if not os.path.exists(folder_path):
                    self.log_signal.emit(f"'{folder_name}' 文件夹不存在，任务停止")
                    self.result_signal.emit((task_name, ocr_results))
                    return

                ocr_results[img_name] = []

                image_files = natsorted(os.listdir(folder_path))

                for image_filename in image_files:
                    if self.stop_flag['stop']:
                        self.log_signal.emit("OCR任务已停止")
                        self.result_signal.emit((task_name, ocr_results))
                        return
                    img_path = os.path.join(folder_path, image_filename)
                    with open(img_path, "rb") as image_file:
                        image_data = image_file.read()

                    result = ""  # Initialize result
                    if model == "百度OCR(联网)":
                        image_base64 = base64.b64encode(image_data).decode('utf-8')
                        image_base64_urlencoded = urllib.parse.quote(image_base64, safe='')

                        payload = f'image={image_base64_urlencoded}&detect_direction=false&detect_language=false&paragraph=false&probability=false'
                        headers = {
                            'Content-Type': 'application/x-www-form-urlencoded',
                            'Accept': 'application/json'
                        }

                        retry = True
                        while retry:
                            if self.stop_flag['stop']:
                                self.log_signal.emit("OCR任务已停止")
                                return
                            try:
                                response = requests.request("POST", url, headers=headers, data=payload)

                                if response.status_code == 200:
                                    result_json = response.json()
                                    error_code = result_json.get("error_code")

                                    if error_code == 18:
                                        self.log_signal.emit("触发限流，2秒后再次识别")
                                        time.sleep(2)
                                        retry = True
                                    elif error_code == 17:
                                        result = "识别次数不足"
                                        retry = False
                                    else:
                                        if "words_result" in result_json and result_json["words_result"]:
                                            result = "".join([item["words"] for item in result_json["words_result"]])
                                        else:
                                            result = "识别失败"
                                        retry = False
                                else:
                                    result = "识别失败"
                                    retry = False
                            except RequestException:
                                self.log_signal.emit("网络连接失败")
                                self.stop_flag['stop'] = True
                                return

                    elif model == "云端OCR(联网)":
                        image_base64 = base64.b64encode(image_data).decode('utf-8')

                        retry = True
                        while retry:
                            if self.stop_flag['stop']:
                                self.log_signal.emit("OCR任务已停止")
                                return
                            try:
                                response = requests.post(
                                    "https://api.npcbug.com/ocr/tyocr.php",
                                    data={'base64': image_base64, 'auth': auth_code}
                                )

                                if response.status_code == 200:
                                    result_json = response.json()

                                    if result_json.get("error_code") == 403:
                                        self.log_signal.emit("授权码不存在或可用次数不足")
                                        self.stop_flag['stop'] = True
                                        return

                                    if result_json.get("error_code") == 18:
                                        self.log_signal.emit("触发限流，2秒后再次识别")
                                        time.sleep(2)
                                        retry = True
                                    elif result_json.get("error_code") == 17:
                                        result = "识别次数不足"
                                        retry = False
                                    else:
                                        if "words_result" in result_json and result_json["words_result"]:
                                            result = "".join([item["words"] for item in result_json["words_result"]])
                                        else:
                                            result = "识别失败"
                                        retry = False
                                else:
                                    result = "识别失败"
                                    retry = False
                            except RequestException:
                                self.log_signal.emit("网络连接失败")
                                self.stop_flag['stop'] = True
                                return

                    else:
                        image = cv2.imdecode(np.fromfile(img_path, dtype=np.uint8), cv2.IMREAD_GRAYSCALE)
                        if image is None:
                            result = "识别失败"
                            ocr_results[img_name].append((image_filename, result))
                            continue

                        if model == "Ddddocr模型":
                            result = self.ocr_model.classification(image_data)
                        elif model == "Paddle模型":
                            paddle_result = self.ocr_model.ocr(img_path, cls=True)
                            if paddle_result and paddle_result[0]:
                                result = ''.join([line[1][0] for line in paddle_result[0]])
                            else:
                                result = "识别失败"

                    if not result:
                        result = "识别失败"

                    result = replace_text_in_result(result, replace_option, replace_text, self.log_signal)
                    ocr_results[img_name].append((image_filename, result))
                    self.log_signal.emit(f"图片 {image_filename} 的识别结果: {result}")

                    self.image_visualization_signal.emit(task_name, img_path)

                    # 根据模型设置不同的延迟时间
                    if model == "百度OCR(联网)":
                        time.sleep(1.5)  # 1.5秒延迟
                    elif model == "云端OCR(联网)" or model == "Ddddocr模型" or model == "Paddle模型":
                        time.sleep(0.2)  # 200ms延迟

                    step_count += 1
                    self.progress_signal.emit(int((step_count / total_steps) * 100))

                self.log_signal.emit(f"识别 '{img_name}' 结束")

            self.result_signal.emit((task_name, ocr_results))
            self.progress_signal.emit(100)
            self.log_signal.emit("任务结束")

        except Exception as e:
            self.log_signal.emit(f"处理过程中出现错误: {e}\n{traceback.format_exc()}")

        finally:
            self.release_model()

    def initialize_model(self):
        model = self.task['model']
        if model == "Ddddocr模型":
            self.ocr_model = ddddocr.DdddOcr()
        elif model == "Paddle模型":
            model_dir = os.path.join('./models/whl/rec/ch/ch_PP-OCRv4_rec_infer/')
            if not os.path.exists(model_dir):
                self.log_signal.emit(f"PaddleOCR模型目录不存在: {model_dir}")
                self.stop_flag['stop'] = True
                return
            self.ocr_model = PaddleOCR(
                det_model_dir=os.path.join(model_dir, 'det'),
                rec_model_dir=os.path.join(model_dir, 'rec'),
                cls_model_dir=os.path.join(model_dir, 'cls'),
                lang='ch',
                use_angle_cls=True,
                use_gpu=False
            )

    def release_model(self):
        if self.ocr_model:
            del self.ocr_model
            self.ocr_model = None

    def stop(self):
        self.stop_flag['stop'] = True
        self.log_signal.emit("任务手动停止")
        self.release_model()  # 确保手动停止时释放模型

    def export_results(self, task_name, ocr_results):
        try:
            if self.task['export_option'] == "不导出表格":
                self.log_signal.emit(f"{task_name}任务完成，不导出表格")
                return

            if not os.path.exists('excel'):
                os.makedirs('excel')
            output_path = os.path.join('excel', f"{task_name}_识别结果.xlsx")

            if self.is_file_open(output_path):
                self.log_signal.emit(f"请关闭 {output_path} 文件后再次执行任务")
                return

            workbook = Workbook()

            # 获取任务相关的信息
            model = self.task['model']
            replace_option = self.task['replace_option']
            replace_text = self.task['replace_text']
            frame_interval = int(self.task['frame_interval'])
            grayscale_option = self.task['grayscale_option']
            schedule_datetime = self.task['schedule_datetime']  # 新增定时执行信息
            file_path = self.task['file_path']
            image_time_record = self.task.get('image_time_record', "不记录图片时间")  # 获取新的任务参数

            # 从文件名中提取基准时间
            base_time = self.parse_datetime_from_filename(os.path.basename(file_path))
            if base_time is None:
                self.log_signal.emit("无法从文件名中解析日期时间，导出结果时时间点将显示为 '获取失败'")

            for img_name, results in ocr_results.items():
                sheet_name = img_name.split('.')[0]
                worksheet = workbook.create_sheet(title=sheet_name)

                # 调整表头，移除“媒体时间点”，并根据image_time_record设置“图片时间点”
                headers = [
                    "图片文件", "识别结果", "原图",
                    "图片时间点",  # 仅保留“图片时间点”
                    "图片秒数点",
                    "识别模型", "替换选项", "替换文本", "提取间隔", "灰度图像选项", "定时执行"
                ]
                worksheet.append(headers)

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 创建对齐对象，设置水平和垂直居中
                center_alignment = Alignment(horizontal='center', vertical='center')

                for r_idx, (image_file, result) in enumerate(results, start=2):
                    worksheet.cell(row=r_idx, column=1, value=image_file)
                    worksheet.cell(row=r_idx, column=2, value=result)

                    # 提取图片的秒数点，假设文件名中包含了秒数信息
                    try:
                        # 从文件名中提取秒数点，例如文件名格式为 "taskname_foldername_10.png"
                        # 其中的 "10" 是秒数点
                        second_point = int(image_file.split('_')[-1].split('.')[0])
                    except ValueError:
                        second_point = "未知秒数"

                    worksheet.cell(row=r_idx, column=5, value=second_point)

                    # 在每一行中加入这些信息
                    worksheet.cell(row=r_idx, column=6, value=model)
                    worksheet.cell(row=r_idx, column=7, value=replace_option)
                    worksheet.cell(row=r_idx, column=8, value=replace_text)
                    worksheet.cell(row=r_idx, column=9, value=frame_interval)
                    worksheet.cell(row=r_idx, column=10, value=grayscale_option)

                    # 根据schedule_datetime是否为“无需定时”填写
                    if schedule_datetime == "无需定时":
                        worksheet.cell(row=r_idx, column=11, value="无需定时")
                    else:
                        worksheet.cell(row=r_idx, column=11, value=schedule_datetime)  # 保留原有定时执行信息

                    # 根据image_time_record决定“图片时间点”填充内容
                    if image_time_record == "记录图片时间":
                        image_time_point = self.calculate_image_time(base_time, image_file)
                        worksheet.cell(row=r_idx, column=4, value=image_time_point)  # 设置“图片时间点”
                    else:
                        worksheet.cell(row=r_idx, column=4, value="不记录图片时间")

                    img_path = os.path.join('Imagestorage', task_name, sheet_name, image_file)
                    img = OpenpyxlImage(img_path)

                    img_width, img_height = img.width, img.height
                    col_letter = get_column_letter(3)
                    worksheet.column_dimensions[col_letter].width = img_width * 0.14
                    worksheet.row_dimensions[r_idx].height = img_height * 0.75

                    img.anchor = f'C{r_idx}'
                    worksheet.add_image(img)

                    for col in range(1, 12):  # 更新范围以包含新列
                        cell = worksheet.cell(row=r_idx, column=col)
                        cell.border = thin_border
                        cell.alignment = center_alignment  # 设置对齐方式

            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            workbook.save(output_path)
            self.log_signal.emit(f"识别结果已导出至: {output_path}")
        except Exception as e:
            self.log_signal.emit(f"导出结果时出现错误: {e}\n{traceback.format_exc()}")

    def calculate_image_time(self, base_time, image_file):
        if base_time is None:
            # 如果基准时间获取失败，则返回“获取失败”
            return "获取失败"
        try:
            parts = image_file.split('_')
            actual_second = int(parts[-1].split('.')[0])

            image_time_point = base_time + timedelta(seconds=actual_second)

            return image_time_point.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            self.log_signal.emit(f"计算图片时间点时出现错误: {e}")
            return "时间计算错误"

    def is_file_open(self, file_path):
        if os.path.exists(file_path):
            try:
                os.rename(file_path, file_path)
                return False
            except OSError:
                return True
        return False

    def _delete_folder(self, folder_path):
        try:
            shutil.rmtree(folder_path)
        except Exception as e:
            self.log_signal.emit(f"删除文件夹失败: {e}")

    def parse_datetime_from_filename(self, filename):
        """
        从文件名中解析日期时间信息。根据提供的文件名格式，假设日期时间部分为 "YYYYMMDD-HHMMSS"
        例如:
            - 565-20250112-113619.mp4
            - 20250112-113619.mp4
            - 的深V大V-fds-565-20250112-113619.mp4
        """
        try:
            # 使用正则表达式匹配 "YYYYMMDD-HHMMSS" 格式
            match = re.search(r'(\d{4})(\d{2})(\d{2})-(\d{2})(\d{2})(\d{2})', filename)
            if match:
                year, month, day, hour, minute, second = map(int, match.groups())
                return datetime(year, month, day, hour, minute, second)
            else:
                return None
        except Exception as e:
            self.log_signal.emit(f"解析文件名日期时间时出错: {e}")
            return None


class TaskListInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName('TaskListInterface')

        self.main_layout = QVBoxLayout(self)
        self.main_layout.setSpacing(10)
        self.main_layout.setContentsMargins(10, 10, 10, 10)

        self.table_widget = TableWidget(self)

        self.table_widget.setBorderVisible(True)
        self.table_widget.setBorderRadius(8)

        self.table_widget.setWordWrap(False)
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(11)  # 修改为11列，包括“图片时间点”
        self.table_widget.setHorizontalHeaderLabels(
            ['任务名', '识别模型', '导出选项', '视频路径', '识别图片路径', '替换选项', '替换文本', '提取间隔',
             '灰度图像选项', '定时执行', '图片时间点'])  # 移除“媒体时间点”列
        self.table_widget.verticalHeader().hide()

        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.setColumnWidth(0, 100)
        self.table_widget.setColumnWidth(1, 130)
        self.table_widget.setColumnWidth(2, 110)
        self.table_widget.setColumnWidth(3, 120)
        self.table_widget.setColumnWidth(4, 120)
        self.table_widget.setColumnWidth(5, 105)
        self.table_widget.setColumnWidth(6, 100)
        self.table_widget.setColumnWidth(7, 100)
        self.table_widget.setColumnWidth(8, 210)
        self.table_widget.setColumnWidth(9, 200)
        self.table_widget.setColumnWidth(10, 135)

        self.main_layout.addWidget(self.table_widget)

        self.progress_ring = ProgressRing(self)
        self.progress_ring.setTextVisible(True)
        self.progress_ring.setFormat("0%")
        self.progress_ring.setStrokeWidth(10)
        self.progress_ring.setFixedSize(130, 130)

        self.start_button = PushButton('开始执行', self)
        self.start_button.clicked.connect(self.start_task)
        self.stop_button = PushButton('停止任务', self)
        self.stop_button.clicked.connect(self.stop_task)
        self.delete_button = PushButton('删除任务', self)
        self.delete_button.clicked.connect(self.delete_selected_row)
        self.clear_log_button = PushButton('清空日志', self)
        self.clear_log_button.clicked.connect(self.clear_log)

        self.log_text = TextEdit(self)
        self.log_text.setReadOnly(True)

        # 调整 QLabel 的大小，并美化边框
        self.image_label = QLabel(self)
        self.image_label.setFixedSize(150, 130)
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setStyleSheet("""
            border: 1.3px solid #0078D7;
            border-radius: 5px;
            background-color: #f5f5f5;
        """)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.delete_button)
        button_layout.addWidget(self.stop_button)
        button_layout.addWidget(self.clear_log_button)

        control_layout = QVBoxLayout()
        control_layout.setSpacing(5)
        control_layout.addWidget(self.start_button)
        control_layout.addLayout(button_layout)

        progress_button_layout = QHBoxLayout()
        progress_button_layout.addSpacerItem(
            QSpacerItem(12, 0, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))
        progress_button_layout.addWidget(self.progress_ring)
        progress_button_layout.addSpacerItem(
            QSpacerItem(10, 0, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))
        progress_button_layout.addLayout(control_layout)
        progress_button_layout.addSpacerItem(
            QSpacerItem(10, 0, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        progress_log_layout = QHBoxLayout()
        progress_log_layout.addLayout(progress_button_layout)
        progress_log_layout.addWidget(self.log_text)

        progress_log_layout.addSpacerItem(
            QSpacerItem(10, 0, QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Minimum)
        )

        progress_log_layout.addWidget(self.image_label)

        progress_log_layout.addSpacerItem(
            QSpacerItem(12, 0, QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Minimum)
        )

        self.main_layout.addSpacerItem(
            QSpacerItem(0, 12, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))
        self.main_layout.addLayout(progress_log_layout)
        self.main_layout.addSpacerItem(
            QSpacerItem(0, 10, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        self.main_layout.setStretch(0, 5)
        self.main_layout.setStretch(2, 1)

        self.table_widget.itemSelectionChanged.connect(self.on_task_selected)

        self.threads = {}
        self.stop_flags = {}
        self.task_logs = {}
        self.task_progress = {}
        self.task_results = {}
        self.task_images = {}
        self.timers = {}  # 新增：用于存储定时器

    def add_task(self, task_name, model, export_option, file_path, replace_option, image_path, replace_text,
                frame_interval, grayscale_option, schedule_datetime, image_time_record):
        row_position = self.table_widget.rowCount()
        self.table_widget.insertRow(row_position)

        self.table_widget.setItem(row_position, 0, self.create_non_editable_item(task_name))
        self.table_widget.setItem(row_position, 1, self.create_non_editable_item(model))
        self.table_widget.setItem(row_position, 2, self.create_non_editable_item(export_option))
        self.table_widget.setItem(row_position, 3, self.create_non_editable_item(file_path))
        self.table_widget.setItem(row_position, 4, self.create_non_editable_item(image_path))
        self.table_widget.setItem(row_position, 5, self.create_non_editable_item(replace_option))
        self.table_widget.setItem(row_position, 6, self.create_non_editable_item(replace_text))
        self.table_widget.setItem(row_position, 7, self.create_non_editable_item(frame_interval))
        self.table_widget.setItem(row_position, 8, self.create_non_editable_item(grayscale_option))
        self.table_widget.setItem(row_position, 9, self.create_non_editable_item(schedule_datetime))
        self.table_widget.setItem(row_position, 10, self.create_non_editable_item(image_time_record))  # 添加新列

        self.task_logs[task_name] = []
        self.task_progress[task_name] = 0
        self.task_results[task_name] = None
        self.task_images[task_name] = []

        # 默认选中添加的任务
        self.table_widget.selectRow(row_position)

        # 新增功能1：设置定时任务
        if schedule_datetime != "无需定时":
            try:
                scheduled_time = datetime.strptime(schedule_datetime, "%H时%M分").time()
                current_datetime = datetime.now()
                scheduled_datetime = datetime.combine(current_datetime.date(), scheduled_time)
                if scheduled_datetime <= current_datetime:
                    scheduled_datetime += timedelta(days=1)  # 设置为第二天

                delay = (scheduled_datetime - current_datetime).total_seconds()
                if delay > 0:
                    timer = QTimer(self)
                    timer.setSingleShot(True)
                    timer.timeout.connect(lambda tn=task_name: self.execute_scheduled_task(tn))
                    timer.start(int(delay * 1000))  # QTimer以毫秒为单位
                    self.timers[task_name] = timer
                    self.update_log(task_name, f"任务已设置 {schedule_datetime} 定时执行")
                else:
                    # 时间已过，不设置定时器
                    self.update_log(task_name, f"任务的定时时间已过，保持手动执行")
            except ValueError:
                self.update_log(task_name, f"任务的定时时间格式错误，保持手动执行")

    def create_non_editable_item(self, text):
        item = QTableWidgetItem(text)
        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        return item

    def show_info_bar(self, message, info_type='info'):
        if info_type == 'error':
            InfoBar.error(
                title='提示',
                content=message,
                isClosable=True,
                duration=2000,
                parent=self,
                position=InfoBarPosition.TOP_RIGHT
            )
        elif info_type == 'success':
            InfoBar.success(
                title='提示',
                content=message,
                isClosable=True,
                duration=2000,
                parent=self,
                position=InfoBarPosition.TOP_RIGHT
            )
        elif info_type == 'info':
            InfoBar.warning(
                title='提示',
                content=message,
                isClosable=True,
                duration=2000,
                parent=self,
                position=InfoBarPosition.TOP_RIGHT
            )

    def delete_selected_row(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            self.show_info_bar('请选择需要删除的任务', 'info')
            return

        selected_row = self.table_widget.row(selected_items[0])
        task_name = self.table_widget.item(selected_row, 0).text()
        self.table_widget.removeRow(selected_row)
        self.update_log(task_name, f"任务删除成功: {task_name}")
        self.show_info_bar('删除成功', 'success')

        # 重置进度条
        self.progress_ring.setValue(0)
        self.progress_ring.setFormat("0%")

        # 清除日志（仅清除当前显示的日志）
        self.log_text.clear()
        self.image_label.clear()

        if task_name in self.task_logs:
            del self.task_logs[task_name]
        if task_name in self.task_progress:
            del self.task_progress[task_name]
        if task_name in self.threads:
            del self.threads[task_name]
        if task_name in self.task_results:
            del self.task_results[task_name]
        if task_name in self.stop_flags:
            del self.stop_flags[task_name]
        if task_name in self.task_images:
            del self.task_images[task_name]

        # 新增功能2：删除任务时移除定时器
        if task_name in self.timers:
            self.timers[task_name].stop()
            del self.timers[task_name]
            self.update_log(task_name, f"已移除任务 '{task_name}' 的定时器")

        self.delete_task_files(task_name)

        # 自动选中下一个任务（如果有）
        if self.table_widget.rowCount() > 0:
            next_row = selected_row if selected_row < self.table_widget.rowCount() else self.table_widget.rowCount() - 1
            self.table_widget.selectRow(next_row)
            # 手动触发选中项变化的信号
            self.on_task_selected()
        else:
            self.progress_ring.setValue(0)
            self.log_text.clear()
            self.image_label.clear()

    def delete_task_files(self, task_name):
        excel_file = os.path.join('excel', f"{task_name}_识别结果.xlsx")
        if os.path.exists(excel_file):
            os.remove(excel_file)

        identifyimages_dir = os.path.join('Identifyimages', task_name)
        if os.path.exists(identifyimages_dir):
            self._delete_folder(identifyimages_dir)

        imagestorage_dir = os.path.join('Imagestorage', task_name)
        if os.path.exists(imagestorage_dir):
            self._delete_folder(imagestorage_dir)

    def _delete_folder(self, folder_path):
        try:
            shutil.rmtree(folder_path)
        except Exception as e:
            self.log_text.append(f"删除文件夹失败: {e}")
            self.show_info_bar(f"删除文件夹失败: {e}", 'error')

    def start_task(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            self.show_info_bar('请选择需要执行的任务', 'info')
            return

        selected_row = self.table_widget.row(selected_items[0])
        task_name = self.table_widget.item(selected_row, 0).text()

        task = {
            'task_name': task_name,
            'model': self.table_widget.item(selected_row, 1).text(),
            'export_option': self.table_widget.item(selected_row, 2).text(),
            'file_path': self.table_widget.item(selected_row, 3).text(),
            'recognition_image_path': self.table_widget.item(selected_row, 4).text(),
            'replace_option': self.table_widget.item(selected_row, 5).text(),
            'replace_text': self.table_widget.item(selected_row, 6).text(),
            'frame_interval': self.table_widget.item(selected_row, 7).text(),
            'grayscale_option': self.table_widget.item(selected_row, 8).text(),
            'schedule_datetime': self.table_widget.item(selected_row, 9).text(),  # 获取定时执行信息
            'image_time_record': self.table_widget.item(selected_row, 10).text()  # 获取图片时间点信息
        }

        self.update_log(task_name, f"开始执行任务: {task['task_name']}")

        if task_name in self.threads and self.threads[task_name].isRunning():
            self.show_info_bar('任务已在执行中', 'info')
            return

        stop_flag = {'stop': False}
        processing_thread = ImageProcessingThread(task, stop_flag)
        processing_thread.log_signal.connect(lambda msg, t=task_name: self.update_log(t, msg))
        processing_thread.progress_signal.connect(lambda value, t=task_name: self.update_progress(t, value))
        processing_thread.task_complete_signal.connect(lambda t=task_name: self.on_task_complete(t))
        processing_thread.result_signal.connect(self.handle_ocr_results)
        processing_thread.image_visualization_signal.connect(self.store_image_for_task)

        self.threads[task_name] = processing_thread
        self.stop_flags[task_name] = stop_flag
        processing_thread.start()

        self.set_task_row_color(selected_row, QColor("#f3d6ac"))

        # 新增功能2：停止任务时移除定时器
        if task_name in self.timers:
            self.timers[task_name].stop()
            del self.timers[task_name]
            self.update_log(task_name, f"执行任务 '{task_name}' 后已移除定时器")

    def stop_task(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            self.show_info_bar('请选择任务', 'info')
            return

        selected_row = self.table_widget.row(selected_items[0])
        task_name = self.table_widget.item(selected_row, 0).text()

        if task_name not in self.threads or not self.threads[task_name].isRunning():
            self.show_info_bar('任务未在执行过程中', 'info')
            return

        self.stop_flags[task_name]['stop'] = True
        self.show_info_bar('任务停止成功', 'success')

        # 设置任务行颜色为#FA8072
        self.set_task_row_color(selected_row, QColor("#FA8072"))

        # 仅在OCR过程中弹出导出对话框
        if self.threads[task_name].ocr_running:
            self.show_export_dialog(task_name)

        # 新增功能2：停止任务时移除定时器
        if task_name in self.timers:
            self.timers[task_name].stop()
            del self.timers[task_name]
            self.update_log(task_name, f"任务 '{task_name}' 执行中已移除定时器")

    def show_export_dialog(self, task_name):
        w = MessageBox(
            '说明',
            '需要导出已识别的结果吗？',
            self
        )
        w.yesButton.setText('导出')
        w.cancelButton.setText('不导出')
        if w.exec():
            if task_name in self.task_results:
                self.threads[task_name].export_results(task_name, self.task_results[task_name])
        else:
            self.show_info_bar('结果未导出', 'info')

    def handle_ocr_results(self, result):
        task_name, ocr_results = result
        self.task_results[task_name] = ocr_results
        if not self.stop_flags[task_name]['stop']:
            self.threads[task_name].export_results(task_name, ocr_results)

        # 新增功能3：任务执行完毕移除定时器
        if task_name in self.timers:
            self.timers[task_name].stop()
            del self.timers[task_name]
            self.update_log(task_name, f"任务 '{task_name}' 执行完毕，已移除定时器")

    def on_task_selected(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            if self.progress_ring.value() != 0:
                self.progress_ring.setValue(0)
                self.progress_ring.setFormat("0%")
            self.log_text.clear()
            self.image_label.clear()
            return

        task_name = self.table_widget.item(self.table_widget.currentRow(), 0).text()
        self.log_text.clear()
        if task_name in self.task_logs:
            self.log_text.append("\n".join(self.task_logs[task_name]))
        if task_name in self.task_progress:
            self.update_progress(task_name, self.task_progress[task_name])

        self.update_image_display_for_task(task_name)

    def on_task_complete(self, task_name):
        if not self.stop_flags[task_name]['stop']:
            self.show_info_bar(f'{task_name}任务完成', 'success')
            row_position = self.find_task_row(task_name)
            if row_position is not None:
                self.set_task_row_color(row_position, QColor("#c9eabe"))

            # 新增功能3：任务执行完毕移除定时器
            if task_name in self.timers:
                self.timers[task_name].stop()
                del self.timers[task_name]
                self.update_log(task_name, f"任务 '{task_name}' 执行完毕，已移除定时器")

    def find_task_row(self, task_name):
        for row in range(self.table_widget.rowCount()):
            if self.table_widget.item(row, 0).text() == task_name:
                return row
        return None

    def set_task_row_color(self, row, color):
        for col in range(self.table_widget.columnCount()):
            self.table_widget.item(row, col).setBackground(color)

    def update_log(self, task_name, message):
        if task_name not in self.task_logs:
            self.task_logs[task_name] = []
        self.task_logs[task_name].append(message)
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            return
        selected_task_name = self.table_widget.item(self.table_widget.currentRow(), 0).text()
        if selected_task_name == task_name:
            self.log_text.append(message)

    def update_progress(self, task_name, value):
        self.task_progress[task_name] = value
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            return
        selected_task_name = self.table_widget.item(self.table_widget.currentRow(), 0).text()
        if selected_task_name == task_name:
            self.progress_ring.setValue(value)
            self.progress_ring.setFormat(f"{value}%")

    def clear_log(self):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            self.show_info_bar('请选择要清空日志的任务', 'info')
            return

        task_name = self.table_widget.item(self.table_widget.currentRow(), 0).text()
        if task_name in self.task_logs:
            self.task_logs[task_name] = []

        self.log_text.clear()
        self.log_text.append("日志已清空")
        self.show_info_bar('日志已清空', 'success')

    def task_exists(self, task_name):
        for row in range(self.table_widget.rowCount()):
            if self.table_widget.item(row, 0).text() == task_name:
                return True
        return False

    def store_image_for_task(self, task_name, image_path):
        if task_name in self.task_images:
            self.task_images[task_name].append(image_path)

        if self.is_current_task_selected(task_name):
            self.set_image_to_label(image_path)

    def is_current_task_selected(self, task_name):
        selected_items = self.table_widget.selectedItems()
        if not selected_items:
            return False
        selected_task_name = self.table_widget.item(self.table_widget.currentRow(), 0).text()
        return task_name == selected_task_name

    def set_image_to_label(self, image_path):
        pixmap = QPixmap(image_path)
        self.image_label.setPixmap(pixmap.scaled(self.image_label.size(), Qt.AspectRatioMode.KeepAspectRatio))

    def update_image_display_for_task(self, task_name):
        if task_name in self.task_images:
            task_images = self.task_images[task_name]
            if task_images:
                last_image_path = task_images[-1]
                self.set_image_to_label(last_image_path)
            else:
                self.image_label.clear()

    def execute_scheduled_task(self, task_name):
        row_position = self.find_task_row(task_name)
        if row_position is not None:
            # 自动选中任务
            self.table_widget.selectRow(row_position)
            self.start_task()
            self.update_log(task_name, f"任务已按定时时间执行")
            self.show_info_bar(f"任务 '{task_name}' 已按定时时间执行", 'success')